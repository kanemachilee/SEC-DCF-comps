# src/build_dcf_per_company.py
import math
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

WB_PATH = Path("model/valuation_pack.xlsx")
if not WB_PATH.exists():
    raise SystemExit("[ERR] model/valuation_pack.xlsx not found. Run build_comps_and_model.py first.")

# ----------------- Load workbook & base sheet -----------------
wb = load_workbook(WB_PATH)
wsA = wb["Assumptions"]  # baseline Rf/ERP/tax/g

# ----------------- Helpers -----------------
def read_assumption(ws: Worksheet, label: str, default=None):
    for row in ws.iter_rows(min_row=2, max_col=3):
        a = row[0].value
        if a and str(a).strip().lower().startswith(label.lower()):
            try:
                return float(row[1].value)
            except Exception:
                return default
    return default

def read_override(ws: Worksheet, label: str, default=None):
    for r in ws.iter_rows(min_row=1, max_row=40, min_col=1, max_col=2):
        a = r[0].value
        if a and str(a).strip().lower().startswith(label.lower()):
            try:
                return float(r[1].value)
            except Exception:
                return default
    return default

def load_comps():
    df = pd.read_csv("data_proc/comps.csv", encoding="utf-8-sig")
    # normalize headers and ticker
    df.columns = [c.encode("utf-8","ignore").decode("utf-8").strip().lower() for c in df.columns]
    if "ticker" not in df.columns:
        raise SystemExit("[ERR] comps.csv missing 'ticker' column. Rebuild comps.")
    df["ticker"] = df["ticker"].astype(str).str.strip().str.upper()
    # numeric coercion
    for col in ["price","dilutedshares","equityvalue","netdebt","ev","revenue (fy)","ebit (fy)"]:
        if col in df.columns: df[col] = pd.to_numeric(df[col], errors="coerce")
    return df

def scale_shares_if_needed(sh):
    if pd.isna(sh): return sh
    if sh < 10_000:  # looks like "in millions"
        return sh * 1_000_000
    return sh

# ----------------- Baseline assumptions -----------------
Rf = read_assumption(wsA, "Risk-free", default=None)
ERP = read_assumption(wsA, "ERP",      default=5.5)
tax_rate = read_assumption(wsA, "Tax rate", default=25.0)
g_pct = read_assumption(wsA, "Terminal growth", default=2.5)
beta_u = read_assumption(wsA, "Industry beta",  default=0.85)
if Rf is None: raise SystemExit("[ERR] Risk-free (10Y, %) missing in Assumptions.")

Rf_d, ERP_d, tax_d, g_d = Rf/100.0, ERP/100.0, tax_rate/100.0, g_pct/100.0

# ----------------- Data prep -----------------
comps = load_comps()
fin = pd.read_csv("data_proc/financials_tidy.csv", encoding="utf-8-sig")
fin["ticker"] = fin["ticker"].astype(str).str.strip().str.upper()
lastfy = (fin.sort_values(["ticker", "fy"]).groupby("ticker").tail(1).set_index("ticker"))

tickers = comps["ticker"].dropna().astype(str).str.strip().str.upper().unique().tolist()
print("Detected tickers:", tickers)

# Preserve existing overrides if present
existing_overrides = {}
for t in tickers:
    name = f"{t}_DCF"
    if name in wb.sheetnames:
        ws_old = wb[name]
        existing_overrides[t] = {
            "Growth (rev %)":  read_override(ws_old, "Growth (rev %)",  None),
            "EBIT margin (%)": read_override(ws_old, "EBIT margin (%)", None),
            "D&A (% rev)":     read_override(ws_old, "D&A (% rev)",     None),
            "CapEx (% rev)":   read_override(ws_old, "CapEx (% rev)",   None),
            "ΔNWC (% rev)":    read_override(ws_old, "ΔNWC (% rev)",    None),
            "Terminal g (%)":  read_override(ws_old, "Terminal g (%)",  None),
            "Tax rate (%)":    read_override(ws_old, "Tax rate (%)",    None),
        }
        wb.remove(ws_old)
    else:
        existing_overrides[t] = {k: None for k in [
            "Growth (rev %)","EBIT margin (%)","D&A (% rev)","CapEx (% rev)","ΔNWC (% rev)","Terminal g (%)","Tax rate (%)"
        ]}

summary = []

# ----------------- Build each company DCF -----------------
for t in tickers:
    row_c = comps.set_index("ticker").loc[t]

    # Revenue base: prefer tidy; else comps FY; else pack average; else $1B
    if t in lastfy.index and pd.notnull(lastfy.loc[t].get("revenue")):
        revenue_base = float(lastfy.loc[t]["revenue"])
    elif pd.notnull(row_c.get("revenue (fy)")):
        revenue_base = float(row_c["revenue (fy)"])
    else:
        pack_rev = pd.to_numeric(comps["revenue (fy)"], errors="coerce")
        revenue_base = float(pack_rev.dropna().mean()) if pack_rev.notna().any() else 1_000_000_000.0
    if not (np.isfinite(revenue_base) and revenue_base > 1e6):
        revenue_base = 1_000_000_000.0

    # EBIT margin: from comps if sane; else 10%
    rev_fy = float(row_c.get("revenue (fy)")) if pd.notnull(row_c.get("revenue (fy)")) and row_c.get("revenue (fy)") != 0 else 1.0
    ebit_margin = float(row_c.get("ebit (fy)")/rev_fy) if pd.notnull(row_c.get("ebit (fy)")) else 0.10
    if not (0.0 < ebit_margin < 0.30):
        ebit_margin = 0.10

    # Shares: comps -> tidy -> median; then scale if in millions
    if pd.notnull(row_c.get("dilutedshares")):
        shares = float(row_c["dilutedshares"])
    elif t in lastfy.index and pd.notnull(lastfy.loc[t].get("diluted_shares")):
        shares = float(lastfy.loc[t]["diluted_shares"])
    else:
        shares = float(pd.to_numeric(comps["dilutedshares"], errors="coerce").median(skipna=True) or 1.0)
    shares = scale_shares_if_needed(shares)

    # Net Debt: from comps (fallback 0)
    net_debt = float(row_c.get("netdebt")) if pd.notnull(row_c.get("netdebt")) else 0.0

    # D/E from comps approx
    de_ratio = 0.0
    if pd.notnull(row_c.get("netdebt")) and pd.notnull(row_c.get("equityvalue")) and row_c.get("equityvalue") != 0:
        de_ratio = float(row_c["netdebt"] / row_c["equityvalue"])

    # Overrides
    ov = existing_overrides.get(t, {})
    growth_pct = ov.get("Growth (rev %)")  if ov.get("Growth (rev %)")  is not None else 5.0
    ebit_m_pct = ov.get("EBIT margin (%)") if ov.get("EBIT margin (%)") is not None else (ebit_margin * 100.0)
    da_pct     = ov.get("D&A (% rev)")     if ov.get("D&A (% rev)")     is not None else 5.0
    capex_pct  = ov.get("CapEx (% rev)")   if ov.get("CapEx (% rev)")   is not None else 6.0
    nwc_pct    = ov.get("ΔNWC (% rev)")    if ov.get("ΔNWC (% rev)")    is not None else 1.0
    term_g_pct = ov.get("Terminal g (%)")  if ov.get("Terminal g (%)")  is not None else g_pct
    tax_pct    = ov.get("Tax rate (%)")    if ov.get("Tax rate (%)")    is not None else tax_rate

    # To decimals
    growth = growth_pct/100.0
    ebit_margin_use = ebit_m_pct/100.0
    da_d = da_pct/100.0
    capex_d = capex_pct/100.0
    nwc_d = nwc_pct/100.0
    term_g_d = term_g_pct/100.0
    tax_d_local = tax_pct/100.0

    # Relever beta & WACC
    beta_l = beta_u * (1 + (1 - tax_d_local) * de_ratio)
    cost_of_equity = Rf_d + beta_l * (ERP_d)
    WACC = cost_of_equity

    # Build sheet
    ws = wb.create_sheet(f"{t}_DCF")
    ws.append([f"{t} DCF Model"])
    ws.append(["Assumption","Value"])
    ws.append(["Rf (%)", Rf])
    ws.append(["ERP (%)", ERP])
    ws.append(["Unlevered beta", beta_u])
    ws.append(["Levered beta", beta_l])
    ws.append(["Tax rate (%)", tax_pct])
    ws.append(["WACC (%)", WACC*100.0])
    ws.append(["Terminal g (%)", term_g_pct])
    ws.append([])

    ws.append(["Override Inputs (editable in Excel)"])
    ws.append(["Growth (rev %)",    growth_pct])
    ws.append(["EBIT margin (%)",   ebit_m_pct])
    ws.append(["D&A (% rev)",       da_pct])
    ws.append(["CapEx (% rev)",     capex_pct])
    ws.append(["ΔNWC (% rev)",      nwc_pct])
    ws.append(["Terminal g (%)",    term_g_pct])
    ws.append(["Tax rate (%)",      tax_pct])
    ws.append([])

    ws.append(["Year","Revenue (proj)","EBIT","Tax","NOPAT","D&A","CapEx","ΔNWC","FCFF","Discount Factor","PV of FCFF"])

    # FCFF stream (5 years)
    N = 5
    fcff, pv_fcff = [], []
    for i in range(1, N+1):
        rev_i = revenue_base * ((1 + growth)**i)
        ebit_i = rev_i * ebit_margin_use
        tax_i = ebit_i * tax_d_local
        nopat_i = ebit_i - tax_i
        da_i = rev_i * da_d
        capex_i = rev_i * capex_d
        dNWC_i = rev_i * nwc_d
        fcff_i = nopat_i + da_i - capex_i - dNWC_i
        disc = (1 + WACC)**i
        pv_i = fcff_i / disc
        fcff.append(fcff_i); pv_fcff.append(pv_i)
        ws.append([2025 + i, rev_i, ebit_i, tax_i, nopat_i, da_i, capex_i, dNWC_i, fcff_i, 1/disc, pv_i])

    ws.append([])
    eff_wacc = max(WACC, term_g_d + 0.001)
    tv = (fcff[-1] * (1 + term_g_d)) / (eff_wacc - term_g_d)
    pv_tv = tv / ((1 + eff_wacc)**N)
    EV = sum(pv_fcff) + pv_tv
    Equity = EV - (0.0 if math.isnan(net_debt) or math.isinf(net_debt) else net_debt)
    implied = Equity / shares if shares and not math.isnan(shares) else float("nan")

    ws.append(["Terminal Value (PV)", pv_tv])
    ws.append(["Enterprise Value", EV])
    ws.append(["Net Debt", net_debt])
    ws.append(["Equity Value", Equity])
    ws.append(["Implied Price", implied])

    # Sensitivity grid
    ws.append([])
    ws.append([f"Sensitivity: Implied Price ($) — {t}"])
    wacc_pts = [max(0.02, WACC - 0.02), max(0.02, WACC - 0.01), WACC, WACC + 0.01, WACC + 0.02]
    g_pts = [term_g_d - 0.005, term_g_d, term_g_d + 0.005, term_g_d + 0.010]
    ws.append(["g ↓ / WACC →"] + [f"{w*100:.1f}%" for w in wacc_pts])

    def implied_for(w, gterm):
        if not shares or math.isnan(shares) or math.isinf(shares):
            return None
        pv_stream = sum(fcff[j-1] / ((1 + w)**j) for j in range(1, N+1))
        weff = max(w, gterm + 0.001)
        tvv = (fcff[-1] * (1 + gterm)) / (weff - gterm)
        pv_tvv = tvv / ((1 + weff)**N)
        evv = pv_stream + pv_tvv
        eqv = evv - (0 if math.isnan(net_debt) or math.isinf(net_debt) else net_debt)
        val = eqv / shares
        return None if math.isnan(val) or math.isinf(val) else val

    for gval in g_pts:
        row = [f"{gval*100:.1f}%"]
        for w in wacc_pts:
            row.append(implied_for(w, gval))
        ws.append(row)

    summary.append({"ticker": t, "WACC (%)": WACC*100.0, "Terminal g (%)": term_g_pct, "Implied": implied})

# -------- Portfolio summary with market price & upside --------
try:
    latest_px = pd.read_csv("data_proc/latest_prices.csv").set_index("ticker")["last_price"]
    latest_px.index = latest_px.index.astype(str).str.strip().str.upper()
except Exception:
    latest_px = pd.Series(dtype=float)

if "Valuation_Summary" in wb.sheetnames:
    wb.remove(wb["Valuation_Summary"])
wsVS = wb.create_sheet("Valuation_Summary")
wsVS.append(["Ticker","WACC (%)","Terminal g (%)","Implied Price","Market Price","Upside (%)"])
for r in summary:
    t = r["ticker"]
    mkt = float(latest_px.get(t, float("nan"))) if not latest_px.empty else float("nan")
    implied = r["Implied"]
    upside = None
    if mkt and (not math.isnan(mkt)) and (not math.isnan(implied)):
        upside = (implied/mkt - 1) * 100.0
    wsVS.append([t, r["WACC (%)"], r["Terminal g (%)"], implied, mkt, upside])

# Keep simple DCF_Summary too
if "DCF_Summary" in wb.sheetnames:
    wb.remove(wb["DCF_Summary"])
wsS = wb.create_sheet("DCF_Summary")
wsS.append(["Ticker","WACC (%)","Terminal g (%)","Implied Price"])
for r in summary:
    wsS.append([r["ticker"], r["WACC (%)"], r["Terminal g (%)"], r["Implied"]])

print("About to save sheets:", wb.sheetnames)
wb.save(WB_PATH)
print("✅ Rebuilt per-company DCF tabs with interactive inputs + Valuation_Summary")
