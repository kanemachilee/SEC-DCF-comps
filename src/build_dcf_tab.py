# src/build_dcf_tab.py
import math
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# -------- Load workbook & sheets --------
wb_path = Path("model/valuation_pack.xlsx")
if not wb_path.exists():
    raise SystemExit("[ERR] model/valuation_pack.xlsx not found. Run build_comps_and_model.py first.")

wb = load_workbook(wb_path)
wsA = wb["Assumptions"]
wsC = wb["Comps"]

# -------- Helpers --------
def read_cell(ws, label):
    """Find first row with Column A beginning with `label` and return Column B as float if present."""
    for row in ws.iter_rows(min_row=2, max_col=3):
        if str(row[0].value).strip().lower().startswith(label.lower()):
            val = row[1].value
            try:
                return float(val) if val not in (None, "") else None
            except Exception:
                return None
    return None

# -------- Pull assumptions from Excel --------
Rf = read_cell(wsA, "Risk-free")        # percent, e.g., 4.01
ERP = read_cell(wsA, "ERP")             # percent, e.g., 5.5
beta_u = read_cell(wsA, "Industry beta")# unlevered
tax_rate = read_cell(wsA, "Tax rate")   # percent
g_pct = read_cell(wsA, "Terminal growth") # percent

# Safe defaults (you can overwrite in Excel later)
if ERP is None: ERP = 5.5
if beta_u is None: beta_u = 0.85
if tax_rate is None: tax_rate = 25.0
if g_pct is None: g_pct = 2.5
if Rf is None: raise SystemExit("[ERR] Risk-free (10Y, %) missing in Assumptions.")

# Convert to decimals for math
Rf_d = Rf / 100.0
ERP_d = ERP / 100.0
tax_d = tax_rate / 100.0
g_d = g_pct / 100.0

# -------- Build comps summary we need (D/E, shares, net debt) --------
comps = pd.read_csv("data_proc/comps.csv")
if comps.empty:
    raise SystemExit("[ERR] data_proc/comps.csv missing or empty. Run build_comps_and_model.py first.")

# D/E using NetDebt/EquityValue is an approximation; good enough for now
comps["D/E"] = comps["NetDebt"] / comps["EquityValue"]
avg_de_ratio = float(comps["D/E"].mean(skipna=True))

# Relever beta & CAPM
beta_l = beta_u * (1 + (1 - tax_d) * avg_de_ratio)
cost_of_equity = Rf_d + beta_l * ERP_d
WACC = cost_of_equity  # assuming minimal debt for simplicity

# We’ll value “the pack” on average to keep scope under control
net_debt = float(comps["NetDebt"].mean(skipna=True))
shares = float(comps["DilutedShares"].mean(skipna=True))

# -------- Build/replace the DCF sheet --------
if "DCF_Model" in wb.sheetnames:
    wb.remove(wb["DCF_Model"])
wsDCF = wb.create_sheet("DCF_Model")

# Header
wsDCF.append(["Year", "Revenue (proj)", "EBIT", f"Tax @{tax_rate:.1f}%", "NOPAT",
              "D&A", "CapEx", "ΔNWC", "FCFF", "Discount Factor", "PV of FCFF"])

# Projection primitives (super simple pack-level model)
rev_base = float(comps["Revenue (FY)"].mean(skipna=True))
# Guard against divide-by-zero
rev_sum = comps["Revenue (FY)"].sum(skipna=True)
ebit_sum = comps["EBIT (FY)"].sum(skipna=True)
ebit_margin = 0.10 if rev_sum == 0 else float(ebit_sum / rev_sum)

da_pct = 0.05   # D&A ≈ 5% of revenue
capex_pct = 0.06 # CapEx ≈ 6% of revenue
wc_pct = 0.01    # ΔNWC ≈ 1% of revenue
growth = 0.05    # 5% revenue CAGR placeholder

# Build 5-year FCFF stream
N = 5
fcff = []
pv_fcff = []
for t in range(1, N+1):
    revenue = rev_base * ((1 + growth) ** t)
    ebit = revenue * ebit_margin
    tax = ebit * tax_d
    nopat = ebit - tax
    da = revenue * da_pct
    capex = revenue * capex_pct
    dNWC = revenue * wc_pct
    fcff_t = nopat + da - capex - dNWC
    disc = (1 + WACC) ** t
    pv = fcff_t / disc
    fcff.append(fcff_t); pv_fcff.append(pv)
    wsDCF.append([2025 + t, revenue, ebit, tax, nopat, da, capex, dNWC, fcff_t, 1/disc, pv])

# Terminal value
fcff_TV = fcff[-1] * (1 + g_d)
# Ensure WACC > g to avoid division by zero (nudge if necessary)
eff_WACC = max(WACC, g_d + 0.0025)
TV = fcff_TV / (eff_WACC - g_d)
PV_TV = TV / ((1 + eff_WACC) ** N)
wsDCF.append(["", "", "", "", "", "", "", "", "Terminal Value (PV)", "", PV_TV])

# Summaries
EV = sum(pv_fcff) + PV_TV
Equity = EV - net_debt
implied = Equity / shares if shares and not math.isnan(shares) else float("nan")
wsDCF.append(["", "", "", "", "", "", "", "", "", "Enterprise Value", EV])
wsDCF.append(["", "", "", "", "", "", "", "", "", "Net Debt", net_debt])
wsDCF.append(["", "", "", "", "", "", "", "", "", "Equity Value", Equity])
wsDCF.append(["", "", "", "", "", "", "", "", "", "Implied Price", implied])

# -------- Sensitivity: Implied Price vs WACC and g --------
wsDCF.append([])  # blank row
wsDCF.append(["Sensitivity: Implied Price ($)"])

# Choose points around current assumptions
wacc_points = [eff_WACC - 0.02, eff_WACC - 0.01, eff_WACC, eff_WACC + 0.01, eff_WACC + 0.02]
g_points = [0.015, 0.020, 0.025, 0.030]  # 1.5%, 2.0%, 2.5%, 3.0%

# Header row: WACC %
wsDCF.append(["g ↓ / WACC →"] + [f"{w*100:.1f}%" for w in wacc_points])

def implied_price_for(wacc, gterm):
    # PV of FCFF stream at this WACC
    pv_stream = sum(fcff[t-1] / ((1 + wacc) ** t) for t in range(1, N+1))
    # Guard to keep denominator positive
    wacc_eff = max(wacc, gterm + 0.001)
    tv = (fcff[-1] * (1 + gterm)) / (wacc_eff - gterm)
    pv_tv = tv / ((1 + wacc_eff) ** N)
    ev = pv_stream + pv_tv
    eq = ev - net_debt
    return eq / shares if shares and not math.isnan(shares) else float("nan")

for gval in g_points:
    row = [f"{gval*100:.1f}%"]
    for w in wacc_points:
        row.append(implied_price_for(w, gval))
    wsDCF.append(row)

# -------- Finish --------
wb.save(wb_path)
print(f"Levered beta={beta_l:.2f}, Cost of equity={cost_of_equity*100:.2f}%")
print(f"Added/updated DCF_Model tab with WACC × g sensitivity in {wb_path.name}")
